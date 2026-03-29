from datetime import date
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

GREEN = colors.HexColor('#22c55e')
WHITE = colors.white
GRAY  = colors.HexColor('#6b7280')
DATE_FMT = '%d/%m/%Y'

def _fmt_date(d):
    return d.strftime(DATE_FMT) if d else '-'

def _pdf_header_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GREEN),
        ('TEXTCOLOR',  (0, 0), (-1, 0), WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 10),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#d1fae5')),
        ('FONTSIZE',   (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])

def generar_reporte_mascotas_pdf(estado=''):
    from core.models import Mascota
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=2*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', fontSize=18, textColor=GREEN,
                                 alignment=TA_CENTER, spaceAfter=12, fontName='Helvetica-Bold')
    elements = [
        Paragraph('Reporte de Mascotas - AdoptPets', title_style),
        Paragraph(f'Generado: {_fmt_date(date.today())}',
                  ParagraphStyle('date', fontSize=10, textColor=GRAY, alignment=TA_RIGHT, spaceAfter=6)),
        Spacer(1, 0.5*cm),
    ]
    qs = Mascota.objects.all()
    if estado:
        try:
            qs = qs.filter(estado_adopcion=estado)
        except Exception:
            pass

    headers = ['ID', 'Nombre', 'Especie', 'Raza', 'Edad', 'Sexo', 'Refugio', 'Estado']
    data = [headers]
    for m in qs:
        data.append([
            str(m.id_mascota), m.nombre, m.especie,
            m.raza or '-',
            f'{m.edad_aproximada} años' if m.edad_aproximada else '-',
            m.sexo,
            m.refugio.nombre_refugio if m.refugio else '-',
            m.estado_adopcion,
        ])
    table = Table(data, repeatRows=1)
    table.setStyle(_pdf_header_style())
    elements += [table, Spacer(1, 0.5*cm),
                 Paragraph(f'<b>Total: {qs.count()}</b>', styles['Normal'])]
    doc.build(elements)
    return buffer.getvalue()

def generar_reporte_adopciones_pdf(fecha_inicio=None, fecha_fin=None):
    from core.models import Adopcion
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=2*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', fontSize=18, textColor=GREEN,
                                 alignment=TA_CENTER, spaceAfter=12, fontName='Helvetica-Bold')
    elements = [Paragraph('Reporte de Adopciones - AdoptPets', title_style), Spacer(1, 0.5*cm)]
    qs = Adopcion.objects.select_related('adoptante', 'mascota', 'mascota__refugio').all()
    if fecha_inicio and fecha_fin:
        qs = qs.filter(fecha_solicitud__range=(fecha_inicio, fecha_fin))

    headers = ['ID', 'Adoptante', 'Mascota', 'Refugio', 'Fecha Solicitud', 'Estado', 'Fecha Aprobación']
    data = [headers]
    for a in qs:
        data.append([
            str(a.id_adopcion),
            f'{a.adoptante.nombre} {a.adoptante.apellido}',
            a.mascota.nombre,
            a.mascota.refugio.nombre_refugio if a.mascota.refugio else '-',
            _fmt_date(a.fecha_solicitud),
            a.estado_adopcion,
            _fmt_date(a.fecha_aprobacion),
        ])
    table = Table(data, repeatRows=1)
    table.setStyle(_pdf_header_style())
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()

def _excel_header(ws, headers):
    fill   = PatternFill(start_color='22c55e', end_color='22c55e', fill_type='solid')
    font   = Font(bold=True, color='FFFFFF')
    align  = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill; cell.font = font
        cell.alignment = align; cell.border = border

def _autosize(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

def generar_reporte_mascotas_excel(estado=''):
    from core.models import Mascota
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mascotas'
    _excel_header(ws, ['ID', 'Nombre', 'Especie', 'Raza', 'Edad', 'Sexo', 'Tamaño', 'Refugio', 'Estado', 'Fecha Ingreso'])
    qs = Mascota.objects.all()
    if estado:
        try:
            qs = qs.filter(estado_adopcion=estado)
        except Exception:
            pass
    for m in qs:
        ws.append([
            m.id_mascota, m.nombre, m.especie,
            m.raza or '-', m.edad_aproximada or 0,
            m.sexo, m.tamano,
            m.refugio.nombre_refugio if m.refugio else '-',
            m.estado_adopcion,
            _fmt_date(m.fecha_ingreso) if m.fecha_ingreso else '-',
        ])
    _autosize(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def generar_reporte_adopciones_excel(fecha_inicio=None, fecha_fin=None):
    from core.models import Adopcion
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Adopciones'
    _excel_header(ws, ['ID', 'Adoptante', 'Email', 'Mascota', 'Refugio', 'Fecha Solicitud', 'Estado', 'Fecha Aprobación'])
    qs = Adopcion.objects.select_related('adoptante', 'mascota', 'mascota__refugio').all()
    if fecha_inicio and fecha_fin:
        qs = qs.filter(fecha_solicitud__range=(fecha_inicio, fecha_fin))
    for a in qs:
        ws.append([
            a.id_adopcion,
            f'{a.adoptante.nombre} {a.adoptante.apellido}',
            a.adoptante.email,
            a.mascota.nombre,
            a.mascota.refugio.nombre_refugio if a.mascota.refugio else '-',
            _fmt_date(a.fecha_solicitud),
            a.estado_adopcion,
            _fmt_date(a.fecha_aprobacion),
        ])
    _autosize(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()